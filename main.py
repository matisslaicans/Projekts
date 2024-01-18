# EXCEL
from openpyxl import Workbook, load_workbook 
# UI
import tkinter as tk
from tkinter import filedialog
# WEB SCRAPING
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# LAIKS
from datetime import datetime
import time

# def open_files():
#     global file_paths
#     file_paths = filedialog.askopenfilename(
#         title='Atver excel vai pdf',
#         filetypes=(("Excel", "*.xlsx"),("All files", "*.*"),("PDF","*.pdf"))
#     )
    
## Excel
file_path='PAMATS.xlsx'
wb=load_workbook(file_path,data_only=True)
ws=wb['4.lapa']
# Platibas
s_pl=round(float(ws['C8'].value),2)-round(float(ws['C24'].value),2)
j_pl=round(float(ws['C14'].value),2)
gr_pl=round(float(ws['C17'].value),2)
kopej_pat=round(float((ws['K14'].value)),2)+round(float((ws['K17'].value)),2)+round(float((ws['K20'].value)),2)+round(float((ws['K24'].value)),2)
wb.close()
## Web Scraping
# Laiks - B, Materials - A, Cena - C, Biezums - D
# Atver excel, kur tiks saglabāti dati
wb=load_workbook('DATA.xlsx')
ws=wb['DATA']
max_row=ws.max_row
# Chrome
service = Service()
option = webdriver.ChromeOptions()
# Getting rid of useless error - most likely USB related
option.add_argument("--log-level=3")
option.add_argument("--disable-logging")
option.add_argument("--log-path=/logs/chromedriver.log")
driver = webdriver.Chrome(service=service, options=option)
# Ieiet majaslapas un accepto cookies
url = "https://prof.lv/"
driver.get(url)
time.sleep(1)
cookies=driver.find_element(By.CLASS_NAME,'ch2-dialog-actions button.ch2-btn-text-sm')
cookies.click()
time.sleep(1)
## Part of the DOES NOT WORK
# url = "https://patatimber.lv/"
# driver.get(url)
# time.sleep(1)
# button=driver.find_element(By.ID,'updateButton')
# button.click()
# time.sleep(0.1)
## Atrod cenu
for i in range(2,max_row+1):
    mater=ws['A'+str(i)].value
    biez=str(ws['D'+str(i)].value)
    x = mater + " " + biez
    
    ## THIS DOES NOT WORK ATM
    # Parbauda vai koks
    # if mater=="Koks":
    #     url = "https://patatimber.lv/"
    #     driver.get(url)
    #     time.sleep(1.5)
    #     # Atrodam cenu priedei pēc biezuma
    #     biez=float(biez)-5
    #     biez=str(biez)
    #     priede=str('priede')
    #     c24s=str('C24')
    #     kd18=str('kd18%')
    #     mekletajs=priede+" "+biez+" "+c24s+" "+kd18
    #     search=driver.find_element(By.ID, 'search-input-field')
    #     search.send_keys(mekletajs)
    #     time.sleep(0.5)
    #     poga = driver.find_element(By.XPATH, '//input[@type="submit" and @value="Meklēt"]')
    #     poga.click()
    #     time.sleep(0.5)
    #     koks=driver.find_element(By.XPATH, f'//a[contains(@title, "{biez}") and contains(@title, "{priede}") and contains(@title, "{c24s}") and contains(@title, "{kd18}")]')
    #     cena=koks.find_element(By.CLASS_NAME,'price')
    #     ws['C'+str(i)].value=cena.text()
    #     laiks=datetime.now()
    #     ws['B'+str(i)].value=laiks
    
    url = "https://prof.lv/"
    driver.get(url)
    time.sleep(1)
    # Atrod search box un samekle no saraksta materialu
    search=driver.find_element(By.CLASS_NAME,'SearchInput')
    search.send_keys(x)
    time.sleep(1)
    find=driver.find_element(By.CLASS_NAME,'action.search')
    find.click()
    time.sleep(3)
    # Cenu samekle - tikai pirmo produktu saraksta, jo otrs ir vesela krava uzreiz
    location=driver.find_element(By.CLASS_NAME,'measure-price-info')
    cena=location.find_element(By.CLASS_NAME,'price')
    ws['C'+str(i)].value=float(cena.text[1:].replace(',','.'))
    laiks=datetime.now()
    ws['B'+str(i)].value=laiks
## Veic apreķinu    
for i in range(2,4):
    ws=wb['DATA']
    bloksm2=float(ws['C'+str(i+2)].value)/(float(ws['D'+str(i+2)].value/1000))
    siltums=float(ws['C'+str(i)].value)
    B_lamb=float(ws['E'+str(i+2)].value)+float(ws['F'+str(i+2)].value)
    S_lamb=float(ws['E'+str(i)].value)+float(ws['F'+str(i)].value)
    B_biez=float(ws['D'+str(i+2)].value)
    S_biez=float(ws['D'+str(i)].value)
    summa=s_pl*(siltums+bloksm2)
    konstrukcija=str(ws['A'+str(i+2)].value)+" "+str(ws['D'+str(i+2)].value)+" + "+str(ws['A'+str(i)].value)+" "+str(ws['D'+str(i)].value)
    Udef=0.23
    Uvert=round(1/(round(((S_biez/1000)/S_lamb),3)+round(((B_biez/1000)/B_lamb),3)+0.13+0.04),2)
    uzlabojums=((kopej_pat+s_pl*Uvert)/(kopej_pat+s_pl*0.23))*100
    per_eur=summa/(100-uzlabojums)
    ws=wb['RESULT']
    ws['A'+str(i)]=konstrukcija
    ws['B'+str(i)]=summa
    ws['C'+str(i)]=Uvert
    ws['D'+str(i)]=uzlabojums
    ws['E'+str(i)]=per_eur
wb.save('DATA.xlsx')
wb.close()



## 1. Izdomāt 3 konstrukcijas sienai (visām vienādi U)             --- DARĪTS (2)
# gāzbetons 300 + 200mm?    0.12
# koka karkasa 200 + 100 + 25 koksiedra abas puses
# ecoterm + 150mm
## Add extra Jumtam?
# Pārsegumu
# Slīpo

## 2. Iegūt nepieciešamās platību vērtības no Excel             --- DARĪTS
# Sakumā tikt pie sienas platībām.
## Optional
# Ņemt visas platības - aprēķināt materiālu izmaksas visai ēkai (dot opciju izvēlēties konstrukcijas)
# Ielikt pārbaudi, vai ir vairāk nekā viena konstrukcija - ja nākamā skaitliskā konstrukcija ir 0, tad vairāk par iepriekšējo skaitlisko vērtību nav.


## 3. Izmantojot web scraping - iegūt aktuālās cenas - konkrētājam kontrukcijām.             --- DARĪTS
## Optional
# Mēģināt pievienot klāt ar laiku vēl mājaslapas, lai var izvelēties, kur materiāli maksā vismazāk.
# Uztaisit pogu "atjaunot datus" - kas izdara tikai parbaudi, vai dati ir atjaunoti ilgak par x


## 4. Izveidot UI
# Norādīt konstrukcijas un to materiālu cenas
# Izveidot pogu, faila izvelei, lai var izvēlēties atsevišķi dažādus projektus.
# ^^^ Kad izvēlēts projekts pēc pārbaudes - ievietot pazīmi, ka ir iespējams pildīt funkcijas.
# Poga - start aprēķins.
# Pielikt cenas ilgumu (cik bieži pārbauda)

## Optional
# Izveidot pārbaudi, pirms cik ilga laika cenas jau ir izpētītas (max 1x stundā)
# Maybe pielikt opciju izvēlēties pdf failu, kur ir jau pabeigtais energosertifikāts un ņemt informāciju no turienes, lai varētu veikt tos pašus apreķinus.







# ## INTERFACE
# # Galvenais logs, ar ko tiek veiktas darbības
# root = tk.Tk()
# # Izmēri
# root.resizable(False, False)
# root.geometry("800x500")
# root.title("Ēku būvniecības izmaksas aprēķins")

# # Pogas
# button1 = tk.Button(root, text="Atvērt failu", font=('Arial',16), command=open_files)
# button1.place(x=10, y=440, height=60, width=200)


# root.mainloop()


